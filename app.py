import openpyxl as xl
from pathlib import Path


def split_onecolumnexelfile(filetosplit, tohowmanycoluns):
    file = Path(filetosplit)
    if file.exists():
        print("Source file  found")
    else:
        print("Source file NOT FOUND")

    wb = xl.load_workbook(filetosplit)
    sheet = wb['Sheet1']

    splitintothismanycolumns = tohowmanycoluns

    split = sheet.max_row // splitintothismanycolumns
    # splitintervalstart = 1
    # splitintervalend = splitintervalstart + split
    amountofrowtodelete = 0
    for i in range(1, splitintothismanycolumns):
        splitintervalstart = i * split
        splitintervalend = splitintervalstart + split
        for row in range(splitintervalstart, splitintervalend):
            sheet.cell(row-(splitintervalstart-1), i+1).value = sheet.cell(row+1, 1).value
            amountofrowtodelete = amountofrowtodelete + 1

    print(amountofrowtodelete)
    sheet.delete_rows(1+split, amount=amountofrowtodelete)
    outputfile = "(v2)" + filetosplit
    wb.save(outputfile)


split_onecolumnexelfile("lotto59-2.xlsx", 3)
